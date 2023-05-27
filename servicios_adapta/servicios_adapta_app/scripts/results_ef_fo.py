import pandas as pd
import openpyxl as xl
import numpy as np


#detectar si el archivo es ef o fo
def file_class(file):
    if ('FO' in file):
        return True
    else:
        return False

def file_info(file):
    file_excel = xl.load_workbook(file, data_only=True, read_only=True)
    file_excel_ws1 = file_excel[file_excel.sheetnames[1]]
    file_excel_ws2 = file_excel[file_excel.sheetnames[2]]


    code = file_excel_ws1['D2'].value
    laeq =  np.round(file_excel_ws1['I5'].value, 2)
    max_val = file_excel_ws1['T5'].value
    min_val = file_excel_ws1['U5'].value
    l10 = file_excel_ws2['C3'].value
    l50 = file_excel_ws2['G3'].value
    l90 = file_excel_ws2['K3'].value
    
    return {'code':code, 'laeq':laeq, 'max_val':max_val, 'min_val': min_val, 'l10':l10, 'l50':l50, 'l90':l90}


def resultados_efs(ef, ws):
    #encontrar código en columna F
    code_row = ws['F']
    code_row_values = []
    for cell in code_row[2:]:
        code_row_values.append(cell.value)

    code_row_values = [str(val).lower().replace(" ", "") for val in code_row_values]
    code_row_values.remove('none')

    i = 0
    code = str(ef['code']).lower().replace(' ', '')
    while (code != code_row_values[i]) and (i < len(code_row_values)-1):
        i += 1

    #asignas valores en fila si existe código
    if (code == code_row_values[i]):
        i += 3
        ws['U'+str(i)].value = ef['laeq']
        ws['X'+str(i)].value = ef['max_val']
        ws['Y'+str(i)].value = ef['min_val']
        ws['Z'+str(i)].value = ef['l10']
        ws['AA'+str(i)].value = ef['l50']
        ws['AB'+str(i)].value = ef['l90']
    
    return ws

def resultados_fos(fo, ws, index):
    frente = fo['code']
    frente = 'R' + frente[1:]
    ws['B' + str(index)].value = frente

    pk = fo['code'].split('_')[-1]
    pk = pk.replace(',', '')
    pk =  int(pk)
    ws['D'+ str(index)].value = pk

    ws['E' + str(index)].value = fo['code']
    ws['N'+str(index)].value = fo['laeq']
    ws['Q'+str(index)].value = fo['max_val']
    ws['R'+str(index)].value = fo['min_val']
    ws['S'+str(index)].value = fo['l10']
    ws['T'+str(index)].value = fo['l50']
    ws['U'+str(index)].value = fo['l90']
    return ws

#todos los procesos juntos
def pasar_resultados_effo(files):
    file_efs_wb = xl.load_workbook('excel_templates/GVC_FCC_R_NPS_EF_MMM_AAAA.xlsx')
    file_efs_ws = file_efs_wb[file_efs_wb.sheetnames[0]]
    file_fos_wb = xl.load_workbook('excel_templates/GVC_FCC_R_NPS_FO_MMM_AAAA.xlsx')
    file_fos_ws = file_fos_wb[file_fos_wb.sheetnames[0]]
    
    fos = []
    efs = []
    for file in files:
        info = file_info(file)
        if file_class(info['code']):
            fos.append(info)
        else:
            efs.append(info)
    
    for ef in efs:
        file_efs_ws = resultados_efs(ef, file_efs_ws)
    fo_index = 3
    for fo in fos:
        file_fos_ws = resultados_fos(fo, file_fos_ws, fo_index)
        fo_index += 1
    
    # file_fos_wb.save('excel_templates/test_fo_results.xlsx')
    return file_efs_wb, file_fos_wb

# file_dir1 = 'excel_templates/02-P_ABR23_FO_4,535.xlsx'
# file_dir2 = 'excel_templates/03-P_ABR23_FO_6,000.xlsx'
# pasar_resultados_effo([file_dir1, file_dir2])