import pandas as pd
import numpy as np
import openpyxl as xl
import os
import csv

#import csv
# def csv_df(path):
#     df = pd.read_csv(path)
#     return df

def detectar_separador_csv(archivo_csv):
    with open(archivo_csv, 'r') as f:
        dialecto = csv.Sniffer().sniff(f.read(1024))
        return dialecto.delimiter

def csv_df(path):
    ruta_temporal = 'archivo.csv'
    with open(ruta_temporal, 'wb') as archivo_temporal:
        for chunk in path.chunks():
            archivo_temporal.write(chunk)


    separador = detectar_separador_csv(ruta_temporal)
    print(separador)
    df = pd.read_csv(ruta_temporal, header=None, sep=separador)
    nombres_columnas = df.iloc[0]
    if nombres_columnas[0] == 'Fecha':
        df.columns = nombres_columnas
        df = df[1:].reset_index(drop=True)
        df = df.apply(lambda x: x.str.replace(",", "."))
        df[df.columns[2]] = df[df.columns[2]].astype(float)
    else:
        nombres_columnas = ['Fecha', 'Tiempo', 'Laeq']
        df.columns = nombres_columnas

    os.remove(ruta_temporal)
    return df

#load template excel
def template_excel(path):
    temp_wb = xl.load_workbook(path)
    return temp_wb

#name code
def name_code_and_date(path , ef):
    if ("\\" in path):
        codigo = path.split("\\")[-1]
    else:
        codigo = path.split("/")[-1]

    codigo = codigo.split(".")[0]
    
    #check if name is ef/fo or not
    if(ef):
        proyect = ""
        if ("FO") not in path:
            if ("_") in codigo:
                date = codigo.split("_")[-1]
                codigo = codigo.split("_")[0]
            else:
                date = ""
        else:
            date = ""
    else:
        date = codigo.split("_")[1]
        proyect = codigo.split("_")[2]
        codigo = codigo.split("_")[0]
    
    return codigo, date, proyect

def create_analysis(csv, template_ws,mins, ef):
    
    #selecting temp_path
    # temp_path = None
    # if (mins == 30):
    #     temp_path = os.path("./noise_templates/template_30.xlsx")
    # elif (mins == 15):
    #     temp_path = os.path("./noise_templates/template_15.xlsx")
    # elif (mins == 60):
    #     temp_path = os.path("./noise_templates/template_60.xlsx")
    
    # if temp_path is not None:
    #     #     #loading
    #     template = template_excel(temp_path)
    #     template_ws = template[template.sheetnames[1]]

    # template_ws = temp[temp.sheetnames[1]]

    data = csv_df(csv)
    print(data.head())
    name_code, date_code, proyect = name_code_and_date(csv.name, ef)

    #columns
    columnas_data = []
    for col in data.columns:
        columnas_data.append(col)
    
    #total time of analysis
    total_mins = int(len(data)/60)

    if (int(mins) < total_mins):
        total_rows = (mins * 60)
    else:
        total_rows = total_mins * 60
        
    data = data[0:total_rows]

    #convert to dateTime else error
    data["Fecha"] = pd.to_datetime(data["Fecha"], dayfirst=True)
    data["Tiempo"] = pd.to_datetime(data["Tiempo"])
    data["Tiempo"] = data["Tiempo"].dt.time

    #inserting data at excel worksheet
    template_ws["A2"].value = data[columnas_data[0]][0]
    template_ws["B2"].value = data[columnas_data[1]][0]
    template_ws["C2"].value = data[columnas_data[1]][total_rows-1]
    template_ws["D2"].value = name_code

    i = 0
    j = 5

    while i < total_rows:
        template_ws['A'+str(j)].value = data[columnas_data[1]][i]
        template_ws['B'+str(j)].value = data[columnas_data[2]][i]
        j += 1
        i += 1

    if (ef) and ('FO' not in name_code):
        save = name_code + "_Operación_" + date_code + ".xlsx"
    elif (ef) and ('FO' in name_code):
        save = name_code + ".xlsx"
    elif (not ef):
        save = name_code + "_" + date_code + "_" + proyect + "_PRNPS" + ".xlsx"

    # template.save(save)
    return template_ws, save
