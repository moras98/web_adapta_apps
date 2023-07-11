from django.db import models
import openpyxl as xl
from datetime import datetime

class Proyecto(models.Model):
    id = models.AutoField(primary_key=True)
    nombre = models.CharField(max_length=100)
    fecha_inicio = models.DateField(auto_now_add=True)

    def __str__(self):
        return self.nombre

class Punto(models.Model):
    id = models.AutoField(primary_key=True)
    proyecto = models.ForeignKey(Proyecto, on_delete=models.CASCADE)
    nombre = models.CharField(max_length=100)

    def __str__(self):
        return f"Punto {self.id} - Proyecto: {self.proyecto.nombre}"

class Medicion(models.Model):
    fecha_inicio = models.DateField()
    punto = models.ForeignKey(Punto, on_delete=models.CASCADE)
    hora_inicio = models.TimeField()
    hora_fin = models.TimeField()
    minutos = models.IntegerField()
    minuto_estabilizacion = models.IntegerField()
    laeq = models.FloatField()
    l10 = models.FloatField()
    l20 = models.FloatField()
    l30 = models.FloatField()
    l40 = models.FloatField()
    l50 = models.FloatField()
    l60 = models.FloatField()
    l70 = models.FloatField()
    l80 = models.FloatField()
    l90 = models.FloatField()
    estandard = models.FloatField()

    def __str__(self):
        return f"Medicion - Fecha: {self.fecha_inicio}, Punto: {self.punto.nombre}"

    def agregar_medicion(cls, excel_file):
        workbook = xl.load_workbook(excel_file, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[1]]
        worksheet2 = workbook[workbook.sheetnames[2]]

        fecha_inicio = worksheet['A2'].value
        if (isinstance(fecha_inicio, str)):
            fecha_inicio = datetime.strptime(fecha_inicio, "%d/%m/%Y")
        punto_nombre = worksheet['D2'].value
        if '0' in punto_nombre:
            punto_nombre = punto_nombre.replace('0', '')
        punto_obj = Punto.objects.get(nombre = punto_nombre)
        punto = punto_obj.id
        hora_inicio = worksheet['B2'].value
        hora_fin = worksheet['C2'].value
        minutos = worksheet['I3'].value
        laeq = worksheet['I5'].value
        l10 = worksheet2['C3'].value
        l20 = worksheet2['D3'].value
        l30 = worksheet2['E3'].value
        l40 = worksheet2['F3'].value
        l50 = worksheet2['G3'].value
        l60 = worksheet2['H3'].value
        l70 = worksheet2['I3'].value
        l80 = worksheet2['J3'].value
        l90 = worksheet2['K3'].value
        estandard = 75

        i = 5
        while (worksheet['R'+str(i)].value == 'No'):
            i += 1

        minuto_estabilizacion = worksheet['L'+str(i)].value

        if not cls.objects.filter(fecha_inicio = fecha_inicio, punto = punto_obj, hora_inicio = hora_inicio).exists():
            medicion = cls(
                fecha_inicio = fecha_inicio,
                punto = punto_obj,
                hora_inicio = hora_inicio,
                hora_fin = hora_fin,
                minutos = minutos,
                minuto_estabilizacion = minuto_estabilizacion,
                laeq = laeq,
                l10 = l10,
                l20 = l20,
                l30 = l30,
                l40 = l40,
                l50 = l50,
                l60 = l60,
                l70 = l70,
                l80 = l80,
                l90 = l90,
                estandard = estandard,
            )
            medicion.save()
        else: return


class experienciaRazonSocial(models.Model):
    id = models.AutoField(primary_key=True)
    nombre = models.CharField(max_length=150, unique=True)
    
    def __str__(self):
        return self.nombre

class experienciaLocalizaciones(models.Model):
    departamento = models.CharField(max_length=100, blank=True)
    pais = models.CharField(max_length=100)

    def __str__(self):
        return f"{self.departamento, self.pais}"


class experienciaProyecto(models.Model):
    id = models.AutoField(primary_key=True)
    nombre = models.CharField(max_length=300)
    contacto_nombre = models.CharField(max_length=100)
    contacto_telefono = models.CharField(max_length=20, blank=True)
    contacto_mail = models.EmailField(blank=True)
    razon = models.ForeignKey(experienciaRazonSocial, on_delete=models.CASCADE)
    localizacion = models.ManyToManyField(experienciaLocalizaciones, related_name='proyectos')
    SECTOR_CHOICES = [
        ('agroindustrial', 'Agroindustrial'),
        ('aguapot_sanea', 'Agua potable y saneamiento'),
        ('ambiente', 'Ambiente'),
        ('comercial', 'Comercial'),
        ('energia', 'Energía'),
        ('industrial', 'Industrial'),
        ('infraestruct', 'Infraestructura'),
        ('logistica', 'Logística'),
        ('mineria', 'Minería'),
        ('puertos', 'Puertos'),
        ('recursosH', 'Recursos Hídricos'),
        ('recursosS', 'Recursos Solidos'),
        ('salud', 'Salud'),
        ('servicios', 'Servicios'),
        ('telecom', 'Telecomunicaciones'),
    ]
    sector = models.CharField(max_length=50, choices=SECTOR_CHOICES)
    
    def __str__(self):
        return self.nombre
    
class experienciaContrato(models.Model):
    fechaInicio = models.DateField()
    

    def fecha_fin_default():#si la borro da error
        pass

    fechaFin = models.CharField(max_length=25, default='En Curso') #Despues se ingresa por el usuario un texto con el formato aaaa-mm-dd
    
    codigo = models.CharField(max_length=8)
    CAT_CHOICES = [
        ('categoria1', 'Categoria 1'),
        ('categoria2', 'Categoria 2'),
        ('categoria3', 'Categoria 3'),
        ('categoria4', 'Categoria 4'),
        ('categoria5', 'Categoria 5'),
        ('categoria6', 'Categoria 6'),
    ]
    catServicios = models.CharField(max_length=50, choices=CAT_CHOICES)
    ficha = models.FileField(null=True, blank=True)
    atestado = models.FileField(null=True, blank=True)
    proyecto = models.ForeignKey(experienciaProyecto, on_delete=models.CASCADE)
    #roles

    def __str__(self):
        return f"Inicio: {self.fechaInicio}, Fin: {self.fechaFin}, RS: {self.proyecto.razon.nombre}, Proyecto: {self.proyecto.nombre}"